"""
This file provides utility functions for working with Excel files stored in Amazon S3, including fetching files,
analyzing sheet structure, and saving output back to S3.
"""

import boto3
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from botocore.exceptions import BotoCoreError, NoCredentialsError, ClientError
from botocore.client import BaseClient
import logging
from types import ModuleType
import re
import importlib
import os
import json
from typing import List
import os

import warnings
from urllib3.exceptions import InsecureRequestWarning

# Suppress before anything else
warnings.simplefilter('ignore', InsecureRequestWarning)

import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def get_secret():
 
    secret_name = "cpce-dev"
    region_name = "us-east-1"
 
    # Create a Secrets Manager client
    session = boto3.session.Session()
    client = session.client(
        service_name='secretsmanager',
        region_name=region_name
    )
 
    try:
        get_secret_value_response = client.get_secret_value(
            SecretId=secret_name
        )
    except ClientError as e:
        # For a list of exceptions thrown, see
        # https://docs.aws.amazon.com/secretsmanager/latest/apireference/API_GetSecretValue.html
        raise e
 
    secret = json.loads(get_secret_value_response['SecretString'])

    return secret

# Boto3 client initialization
def get_s3_client() -> BaseClient:
    """
    Create and return an S3 client using boto3.

    Returns:
        BaseClient : A boto3 S3 client object if successful

    Raise:
        - An error if AWS credentials are not found.
        - An error if there is a BotoCore-related exception while creating the client.
    """
    try:
        # return boto3.client("s3"
        #                     ,verify = False,aws_access_key_id=AWS_ACCESS_KEY_ID, 
        #                         aws_secret_access_key=AWS_ACCESS_KEY_ID, 
        #                         aws_session_token = AWS_SESSION_TOKEN,
        #                         region_name=REGION_NAME)
        return boto3.client("s3"
                            ,verify = False)
    except NoCredentialsError:
        logging.error("AWS credentials not found.")
        raise RuntimeError("AWS credentials not found.")
    except BotoCoreError as e:
        logging.error(f"An error occurred while creating the S3 client: {e}")
        raise RuntimeError(f"Failed to intialize s3 client due to BotocCore error: {e}")


def fetch_excel_from_s3(bucket_name: str, key: str) -> BytesIO:
    """
    Fetch an Excel file from S3 and return its content as a BytesIO object.

    Args:
        bucket_name (str): Name of the S3 bucket.
        key (str): Key (path) of the Excel file in S3.

    Returns:
        BytesIO: In-memory binary stream containing the Excel file content.

    Raises:
        RuntimeError: If there are issues with AWS credentials, S3 access, or file reading.
    """
    try:
        s3 = get_s3_client()
        logger.info(f"Fetching Excel file from S3: Bucket={bucket_name}, Key={key}")
        response = s3.get_object(Bucket=bucket_name, Key=key)

        try:
            file_content = response["Body"].read()
            logger.info(f"Successfully fetched Excel file: {key}")
            return BytesIO(file_content)
        except Exception as e:
            msg = f"Failed to read content of file '{key}': {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

    except ClientError as e:
        error_code = e.response["Error"]["Code"]
        if error_code == "404":
            msg = f"The file at key '{key}' does not exist in bucket '{bucket_name}'."
            logger.error(msg)
            raise RuntimeError(msg) from e
        elif error_code == "403":
            msg = f"Access denied to file '{key}' in bucket '{bucket_name}'."
            logger.error(msg)
            raise RuntimeError(msg) from e
        else:
            msg = f"AWS Error ({error_code}) while fetching '{key}': {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

    except Exception as e:
        msg = f"Unexpected error while fetching Excel file from S3: {e}"
        logger.exception(msg)  # Use exception() to log traceback
        raise RuntimeError(msg) from e


def load_module_from_s3_in_memory(bucket_name: str, file_key: str) -> ModuleType:
    """
    Load a Python module from an S3 object in memory. Module name is derived from the file key.

    Args:
        bucket_name (str): S3 bucket name
        file_key (str): S3 key for the file (e.g., 'path/to/module.py')

    Returns:
        module: The loaded Python module

    Raises:
        RuntimeError: If there are issues fetching, decoding, or executing the module
    """
    try:
        # Derive module name from file_key
        base_name = os.path.basename(file_key)  #  'excel_preview_message.py'
        module_name = os.path.splitext(base_name)[0]  # 'excel_preview_message'

        s3 = get_s3_client()

        try:
            logger.info(
                f"Fetching module code from S3: Bucket={bucket_name}, Key={file_key}"
            )
            # Get the .py file from the s3 bucket for excel-preview or sheet name.
            obj = s3.get_object(Bucket=bucket_name, Key=file_key)
            data = obj["Body"].read().decode("utf-8")
        except Exception as e:
            msg = f"Failed to read or decode file from S3 (Bucket={bucket_name}, Key={file_key}): {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e
        try:
            logger.info(f"Compiling and executing module '{module_name}' in memory.")
            spec = importlib.util.spec_from_loader(module_name, loader=None)
            module = importlib.util.module_from_spec(spec)
            exec(data, module.__dict__)
            return module
        except Exception as e:
            msg = (
                f"Failed to compile or execute module '{module_name}' from S3 code: {e}"
            )
            logger.error(msg)
            raise RuntimeError(msg) from e

    except Exception as e:
        msg = f"[UnexpectedError] Could not load module '{module_name}' from S3: {e}"
        logger.exception(msg)
        raise RuntimeError(msg) from e


def find_matching_sheet_from_s3(
    bucket_name: str,
    key: str,
    sheet_type: str,
    llm_bucket_name_prompt: str,
    file_key: str,
) -> str:
    """
    Find the exact matching sheet name based on sheet type for an Excel file in S3.

    Args:
        bucket_name (str): Name of the S3 bucket.
        key (str): Key (path) of the Excel file in S3.
        sheet_type (str): Type of sheet to find ('budget' or 'assumption' or 'timeline').
        llm_bucket_name_prompt (str): S3 bucket name where the prompts are present
        file_key (str): Key (path) for the Sheet names file

    Returns:
        str: Name of the matching sheet found.

    Raises:
        ValueError: If unsupported sheet type is provided or no matching sheet found.
        RuntimeError: If there's an issue fetching the file from S3 or reading Excel content.
    """
    try:
        # Step 1: Fetch Excel file from S3
        try:
            logger.info(f"Loading Excel file from S3: Bucket={bucket_name}, Key={key}")
            excel_data = fetch_excel_from_s3(bucket_name, key)
        except RuntimeError as e:
            msg = f"Failed to fetch Excel file from S3: {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 2: Read Excel file and get available sheets
        try:
            logger.info("Reading Excel file content.")
            xls = pd.ExcelFile(excel_data)
            available_sheets = xls.sheet_names
        except Exception as e:
            msg = f"Failed to read Excel file content: {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

        try:
            logger.info(
                f"Loading file name '{file_key}' from S3 (Bucket={llm_bucket_name_prompt}, Key={file_key})"
            )
            module = load_module_from_s3_in_memory(llm_bucket_name_prompt, file_key)
            EXPECTED_SHEET_NAMES = module.EXPECTED_SHEET_NAMES
            ASSUMPTIONS_SHEET_NAME = module.ASSUMPTIONS_SHEET_NAME
            TIMELINE_SHEET_NAME = module.TIMELINE_SHEET_NAME

        except Exception as e:
            msg = (
                f"Failed to load or sheet name from S3 "
                f"(Bucket={llm_bucket_name_prompt}, Key={file_key}): {e}"
            )
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 3: Determine expected sheet list based on sheet_type
        if sheet_type == "budget":
            sheet_name_list = EXPECTED_SHEET_NAMES
        elif sheet_type == "assumption":
            sheet_name_list = ASSUMPTIONS_SHEET_NAME
        elif sheet_type == "timeline":
            sheet_name_list = TIMELINE_SHEET_NAME
        else:
            msg = f"Unsupported sheet type: '{sheet_type}'. Expected one of: budget, assumption, timeline."
            logger.error(msg)
            raise ValueError(msg)

        # Step 4: Search for a matching sheet name
        logger.info(f"Searching for matching '{sheet_type}' sheet.")
        for expected_sheet in sheet_name_list:
            matching_sheets = [
                sheet
                for sheet in available_sheets
                if sheet.strip().lower() == expected_sheet.strip().lower()
            ]
            if matching_sheets:
                matched_sheet = matching_sheets[0]
                logger.info(f"Found matching '{sheet_type}' sheet: {matched_sheet}")
                return matched_sheet

        # Step 5: No match found
        msg = f"No matching '{sheet_type}' sheet found. Available sheets: {available_sheets}"
        logger.error(msg)
        raise ValueError(msg)

    except Exception as e:
        msg = f"[UnexpectedError] An unexpected error occurred during processing: {e}"
        logger.exception(msg)
        raise RuntimeError(msg) from e


def metadata_extraction(bucket_name: str, key: str) -> str:
    """
    Fetch metadata JSON content from S3 based on the input Excel file key.
    Derives base name by removing file extension and uses it to locate .xlsx/.xlsm.json files.
    Returns JSON content as string if found, else None.
    """
    logger.info("Extracting metadata from input Excel key.")

    try:
        # Extract base filename without extension
        filename = os.path.basename(key)
        base_key = re.sub(r"\.(xlsx|xlsm)$", "", filename, flags=re.IGNORECASE)
        logger.info(f"Base key extracted: {base_key}")
    except Exception as e:
        msg = f"Error extracting base key from '{key}': {e}"
        logger.error(msg)
        raise RuntimeError(msg) from e

    try:
        s3 = get_s3_client()
    except Exception as e:
        msg = f"Failed to initialize S3 client: {e}"
        logger.error(msg)
        raise RuntimeError(msg) from e

    extensions = ["xlsx", "xlsm"]
    json_content = None

    for ext in extensions:
        json_key = f"metadata/{base_key}_{ext}.json"
        logger.info(f"Trying S3 JSON path: {json_key}")

        try:
            response = s3.get_object(Bucket=bucket_name, Key=json_key)
            logger.info(f"Successfully fetched JSON metadata from S3: {json_key}")
            json_content = response["Body"].read().decode("utf-8")
            break
        except s3.exceptions.NoSuchKey:
            logger.warning(f"JSON metadata not found at: {json_key}")
            continue
        except ClientError as ce:
            msg = f"AWS error fetching JSON from S3 for {json_key}: {ce}"
            logger.error(msg)
            raise RuntimeError(msg) from ce
        except Exception as e:
            msg = f"Unexpected error while fetching JSON from S3 for {json_key}: {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

    if json_content is None:
        msg = "Metadata not found in S3 for either .xlsx or .xlsm."
        logger.warning(msg)

    try:
        metadata = json.loads(json_content)
    except json.JSONDecodeError as e:
        msg = f"Failed to parse metadata JSON: {e}"
        logger.error(msg)
        raise RuntimeError(msg) from e

    # Step 2: Parse metadata fields
    try:
        filename = os.path.basename(key)
        vendor_name = (
            metadata["supplier_rollup_name"][0][0]["label"]
            if metadata.get("supplier_rollup_name")
            else None
        )
        protocol_number = (
            metadata["protocol_study_number"][0]["title"]
            if metadata.get("protocol_study_number")
            else None
        )
        currency = metadata["Currency"]["label"] if metadata.get("Currency") else None
        contract_type = (
            metadata["contract_document"]["label"]
            if metadata.get("contract_document")
            else None
        )
    except Exception as e:
        msg = f"Error parsing metadata fields: {e}"
        logger.error(msg)
        raise RuntimeError(msg) from e

    new_row = {
        "file_name": filename,
        "vendor_name": vendor_name,
        "protocol_number": protocol_number,
        "currency": currency,
        "metadata_contract_type": contract_type,
    }

    return json.dumps(new_row, indent=2)


def get_dimension_headers(sheet_type: str, s3_bucket: str, s3_key: str) -> List[str]:
    """
    Return the appropriate dimension headers based on the sheet type,
    by fetching headers.json from S3.

    Args:
        sheet_type (str): Type of the sheet. Supported values:
                          'budget', 'timeline', 'assumption'
        s3_bucket (str): Name of the S3 bucket where headers.json is located.
        s3_key (str): S3 key/path to headers.json (default: config/headers.json)

    Returns:
        List[str]: List of header names for the specified dimension.

    Raises:
        ValueError: If unsupported sheet type or missing dimension in JSON.
        RuntimeError: If unable to fetch or parse headers.json from S3.
    """
    logger.info(f"Fetching dimension headers from S3: {s3_bucket}/{s3_key}")

    # Map sheet types to their corresponding dimension keys in headers.json
    type_to_dim = {
        "budget": "budget_dim",
        "timeline": "timelines_dim",
        "assumption": "assumptions_dim",
    }

    if sheet_type not in type_to_dim:
        raise ValueError(
            f"Unsupported sheet type: '{sheet_type}'. "
            f"Supported types are: {list(type_to_dim.keys())}"
        )

    dim_key = type_to_dim[sheet_type]

    try:
        s3 = get_s3_client()
        response = s3.get_object(Bucket=s3_bucket, Key=s3_key)
        headers_data = json.loads(response["Body"].read().decode("utf-8"))
        logger.debug("Successfully loaded headers.json from S3")

        if dim_key not in headers_data:
            raise ValueError(
                f"Dimension '{dim_key}' not found in headers.json. "
                f"Available dimensions: {list(headers_data.keys())}"
            )

        return headers_data[dim_key]["headers"]

    except s3.exceptions.NoSuchKey:
        logger.error(f"File not found in S3: {s3_key}")
        raise RuntimeError(f"S3 object not found: {s3_key}")
    except s3.exceptions.ClientError as ce:
        logger.error(f"AWS error fetching headers.json from S3: {ce}")
        raise RuntimeError(f"Failed to fetch headers.json from S3: {ce}")
    except json.JSONDecodeError as je:
        logger.error(f"Invalid JSON format in headers.json: {je}")
        raise RuntimeError(f"Invalid JSON in headers.json: {je}")
    except Exception as e:
        logger.exception("Unexpected error while loading headers from S3:")
        raise RuntimeError(f"Error loading headers from S3: {e}")


def create_complex_excel_preview_from_s3(
    bucket_name: str,
    key: str,
    sheet_name: str,
    sheet_type: str,
    header: str,
    metadata: str,
    llm_bucket_name_prompt: str,
    file_key: str,
) -> str:
    """
    Create a detailed Excel structure preview with comprehensive analysis for an Excel file in S3.

    Args:
        bucket_name (str): Name of the S3 bucket
        key (str): Key (path) of the Excel file in S3
        sheet_name (str): Name of the sheet to analyze
        sheet_type (str): Type of the sheet being analyzed (e.g., 'budget', 'assumption', 'timeline')
        llm_bucket_name_prompt (str): S3 bucket name where the prompts are present
        file_key (str): Key (path) for the excel_preview file

    Returns:
        str: Detailed preview of Excel file structure

    Raises:
        RuntimeError: If there's an issue fetching or analyzing the Excel file
    """
    try:
        # Step 1: Fetch Excel content from S3
        try:
            logger.info(f"Fetching Excel file from S3: Bucket={bucket_name}, Key={key}")
            excel_data = fetch_excel_from_s3(bucket_name, key)
        except Exception as e:
            msg = f"Failed to fetch Excel file from S3: {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 2: Read structured data (with header)
        try:
            df = pd.read_excel(
                excel_data, sheet_name=sheet_name, header=[0, 1], na_values=["NA", ""]
            )
        except Exception as e:
            msg = f"Failed to read Excel sheet '{sheet_name}' with header: {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 3: Reset buffer and read raw data (no header)
        try:
            logger.info(f"Reading raw data for sheet '{sheet_name}'.")
            excel_data.seek(0)
            df_raw = pd.read_excel(
                excel_data, sheet_name=sheet_name, header=None, na_values=["NA", ""]
            )
        except Exception as e:
            msg = f"Failed to read raw Excel sheet '{sheet_name}': {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 4: Analyze workbook-level features using openpyxl
        try:
            logger.info(f"Analyzing workbook features for sheet '{sheet_name}'.")
            excel_data.seek(0)
            wb = load_workbook(filename=BytesIO(excel_data.getvalue()))
            ws = wb[sheet_name]
            merged_ranges = [str(cell_range) for cell_range in ws.merged_cells.ranges]
            contains_formulas = any(
                "=" in str(cell.value) for row in ws.iter_rows() for cell in row
            )
        except Exception as e:
            msg = f"Workbook-level analysis failed for sheet '{sheet_name}': {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 5: Generate dynamic extra info based on sheet type
        try:
            extra_info = ""
            if sheet_type == "timeline":
                logger.info("Adding timeline-specific column explanation.")
                extra_info = (
                    "\n- For the 'Timeline' sheet, headers from the raw file are taken from"
                    f"the values in the first column. \n{df.iloc[:, 1]}"
                )
        except Exception as e:
            msg = f"Failed to generate extra info for sheet '{sheet_name}': {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 5: Load prompt module from S3 and extract preview message template
        try:
            logger.info(
                f"Loading file '{file_key}' from S3 (Bucket={llm_bucket_name_prompt}, Key={file_key})"
            )
            module = load_module_from_s3_in_memory(llm_bucket_name_prompt, file_key)
            excel_preview_message = module.excel_preview_message
        except Exception as e:
            msg = (
                f"Failed to load or parse file '{file_key}' from S3 "
                f"(Bucket={llm_bucket_name_prompt}, Key={file_key}): {e}"
            )
            logger.error(msg)
            raise RuntimeError(msg) from e

        # Step 6: Build final structured message
        try:
            logger.info(f"Formatting final Excel preview for sheet '{sheet_name}'.")

            # Determine how many rows to include in the data sample
            try:
                meta_dict = json.loads(metadata)
                vendor_name = meta_dict.get("vendor_name", "")
            except json.JSONDecodeError:
                logger.warning(
                    "Failed to parse metadata JSON, using default row count."
                )
                vendor_name = ""

            if sheet_type == "assumption" and vendor_name == "IQVIA":
                sample_rows = df_raw.head(30)
            if sheet_type == "timeline" and vendor_name == "IQVIA":
                sample_rows = df_raw.head(30)
            else:
                sample_rows = df_raw.head(15)

            # Capture full string representation without truncation

            data_sample_output = sample_rows.to_string()

            structure_message = excel_preview_message.format(
                file_name=key,
                sheet_name=sheet_name,
                rows=df.shape[0],
                columns=df.shape[1],
                merged_cells_present="Yes" if merged_ranges else "No",
                column_headers=df.columns.to_frame().to_string(),
                data_sample=data_sample_output,  # now either 15 or 30 rows
                num_merged_regions=len(merged_ranges),
                multi_level_headers=(
                    "Yes" if isinstance(df.columns, pd.MultiIndex) else "No"
                ),
                contains_formulas="Yes" if contains_formulas else "No",
                headers=header,
                meta_data_info=metadata,
                extra_info=extra_info,
            )
            logger.info(
                f"Successfully generated Excel preview for sheet '{sheet_name}'."
            )
            return structure_message
        except Exception as e:
            msg = f"Failed to format Excel preview message: {e}"
            logger.error(msg)
            raise RuntimeError(msg) from e

    except RuntimeError as re:
        logger.error(f"[RuntimeError] Error during Excel preview creation: {re}")
        raise
    except Exception as e:
        msg = f"[UnexpectedError] An unexpected error occurred during processing: {e}"
        logger.exception(msg)
        raise RuntimeError(msg) from e


def save_output_to_s3(bucket_name: str, input_key: str, content: str, sheet_name: str) -> str:
    """
    Save the output text content to an S3 bucket with the sheet name in the filename.
    
    Args:
        bucket_name (str): Name of the S3 bucket
        input_key (str): Key (path) of the input file (used to generate the output key)
        content (str): Content to save in the text file
        sheet_name (str): Type of the Excel sheet to include in the output key
    
    Returns:
        str: A success message indicating the S3 object key where the content was saved.
    
    Raises:
        RuntimeError: If there is an error during the process.
    """
    try:
        # Step 1: Get S3 client
        try:
            s3 = get_s3_client()
        except NoCredentialsError as e:
            logger.error("Missing AWS credentials.")
            raise RuntimeError("AWS credentials not found.") from e
        except BotoCoreError as e:
            logger.error("Failed to connect to S3 endpoint.")
            raise RuntimeError("Failed to connect to S3 endpoint.") from e
        except Exception as e:
            logger.error(f"Unexpected error initializing S3 client: {e}")
            raise RuntimeError("Failed to initialize S3 client.") from e

        # Step 2: Generate base output key
        try:
            base_output_key = re.sub(r"\.(xlsx|xlsm)$", "", input_key, flags=re.IGNORECASE)
            output_key = f"{base_output_key}_preview_{sheet_name}.txt"
        except Exception as e:
            logger.error(f"Error generating output key: {e}")
            raise RuntimeError("Error generating output key.") from e

        # Step 3: Upload content
        try:
            logger.info(f"Saving output to S3: Bucket={bucket_name}, Key={output_key}")
            s3.put_object(
                Bucket=bucket_name,
                Key=output_key,
                Body=content.encode("utf-8")
            )
        except ClientError as e:
            logger.error(f"AWS error uploading file: {e}")
            raise RuntimeError("Failed to upload file to S3.") from e
        except Exception as e:
            logger.error(f"Error encoding or uploading content: {e}")
            raise RuntimeError("Error saving content to S3.") from e

        logger.info(f"Output saved successfully to S3: Bucket={bucket_name}, Key={output_key}")
        return f"Output saved successfully to S3"

    except Exception as e:
        msg = f"Unexpected error during output saving to S3: {e}"
        logger.exception(msg)
        raise RuntimeError(msg) from e